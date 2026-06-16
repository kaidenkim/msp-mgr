import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from gcp.constants import DELETABLE_OK, DELETABLE_BILLING, DELETABLE_OWNER


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)


_thin = Side(style="thin", color="BDD7EE")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_HDR   = _fill("1F4E79")
_EVEN  = _fill("DEEAF1")
_ODD   = _fill("FFFFFF")
_GREEN = _fill("E2EFDA")
_RED   = _fill("FCE4D6")
_YLW   = _fill("FFF2CC")


def _summary_sheet(wb, rows):
    ws = wb.create_sheet("요약")
    data = [
        ("항목",                       "수량",    True,  "FFFFFF", "1F4E79"),
        ("전체 프로젝트 (sys- 제외)",  len(rows), False, "000000", "FFFFFF"),
        (DELETABLE_OK,      sum(1 for r in rows if r["deletable"] == DELETABLE_OK),      False, "375623", "E2EFDA"),
        (DELETABLE_BILLING, sum(1 for r in rows if r["deletable"] == DELETABLE_BILLING), False, "833C00", "FCE4D6"),
        (DELETABLE_OWNER,   sum(1 for r in rows if r["deletable"] == DELETABLE_OWNER),   False, "7F6000", "FFF2CC"),
    ]
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    for ri, (label, cnt, bold, fc, bg) in enumerate(data, 1):
        f = _fill(bg)
        for ci, val in [(1, label), (2, cnt)]:
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(bold=bold, color=fc, size=11)
            c.fill = f
            c.border = _border
            c.alignment = Alignment(
                horizontal="center" if ci == 2 else "left", vertical="center"
            )
        ws.row_dimensions[ri].height = 24


def generate_excel(projects: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "GCP 전체 프로젝트"

    headers = [
        "프로젝트 ID", "프로젝트 이름", "생성일",
        "빌링 계정 ID", "빌링 계정 이름", "빌링 계정 상태",
        "소유자(Owner)", "삭제 가능여부",
    ]
    col_widths = [38, 28, 12, 22, 28, 14, 44, 18]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _font(bold=True, color="FFFFFF", size=11)
        c.fill = _HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    ws.row_dimensions[1].height = 28

    for ri, p in enumerate(projects, 2):
        base = _EVEN if ri % 2 == 0 else _ODD
        owners_str = "\n".join(p.get("owners", []))
        n_owners = len(p.get("owners", []))
        b_open = p.get("billing_open", "")
        b_status = "OPEN" if b_open == "True" else ("CLOSED" if b_open == "False" else "")
        row = [
            p["project_id"],
            p["name"],
            p["create_time"],
            p.get("billing_account_id", ""),
            p.get("billing_account_name", ""),
            b_status,
            owners_str,
            p["deletable"],
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border
            if ci == 8:
                d = p["deletable"]
                c.fill = _GREEN if d == DELETABLE_OK else (_RED if d == DELETABLE_BILLING else _YLW)
                c.font = _font(color="375623" if d == DELETABLE_OK else ("833C00" if d == DELETABLE_BILLING else "7F6000"))
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 7:
                c.fill = base; c.font = _font()
                c.alignment = Alignment(wrap_text=True, vertical="top")
            elif ci in (3, 6):
                c.fill = base; c.font = _font()
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = base; c.font = _font()
                c.alignment = Alignment(vertical="center")
        ws.row_dimensions[ri].height = min(max(18, 15 * max(1, n_owners)), 75)

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(projects) + 1}"

    _summary_sheet(wb, projects)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
